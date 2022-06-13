import os
import sys
from pathlib import Path
from openpyxl.utils import get_column_letter
root_path = str(Path(sys.argv[0]).parent)
sys.path.append(root_path)
from myutil.sheet_utils import sheetutil as st
from myutil.time_utils import getDay
from myutil.file_utils import getParentFolder, getFileName
from openpyxl.styles import Font, Fill, Border, Side

def process(sourcefile):
    # 打开一张表
    wb = st.getWorkBook(sourcefile)
    ws = wb.active
    # st.printWB(ws,2)
    # 删掉订购量为0的行
    st.delete_rows_when_equals(ws, '订单量', ['0',0])
    # 删掉不必要的列
    st.deleteByColumnNames(ws, ['商品编码', '要货量', '厂家名称'])
    st.printWB(ws)

    # 删掉后不必要行后，订正第一列的序号
    for i in range(1,ws.max_row):
        ws.cell(i+1, 1).value = str(i)

    # 给最后一行添加日期
    ws.cell(ws.max_row+1, 2).value = getDay()
    # 总数
    # ws.cell(ws.max_row, 3).value = '=SUM(C1:C' + str(ws.max_row-1) + ')'
    # ws['C' + str(ws.max_row)] = '=SUM(C1:C' + str(ws.max_row-1) + ')'
    # ws.cell(ws.max_row, 4).value = '=SUM(D1:D' + str(ws.max_row-1) + ')'
    # 统计信息
    ws['D' + str(ws.max_row)] = '=SUM(D1:D' + str(ws.max_row-1) + ')'
    # ws.cell(ws.max_row, 5).value = '=SUM(E1:E' + str(ws.max_row-1) + ')'
    ws['E' + str(ws.max_row)] = '=SUM(E1:E' + str(ws.max_row-1) + ')'

    # 调整样式
    # 行高
    for i in range(1,ws.max_row):
        ws.row_dimensions[i].height = 22.2

    # 样式-边
    thin = Side(border_style="thin", color="000000")
    b = Border(top=thin, left=thin, right=thin, bottom=thin)
    f = Font(name='Arial', size=16)
    for row in ws.rows:
        for cell in row:
            cell.font = f
            cell.border = b


    # 改列名
    st.rename_firstcolumn(ws, '订单量', '订')

    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '商品')].width = 28
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '订')].width = 4.9
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '批发价')].width = 12
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '金额')].width = 14

    # 保存
    if len(sys.argv) >= 2:
        save_folder = getParentFolder(sys.argv[1]) + "/printer"
    else:
        save_folder = root_path + '/printer'

    if not os.path.isdir(save_folder):
        os.mkdir(save_folder)
    fname = getFileName(sourcefile)
    dest = save_folder + '/' + getDay() + "_" + fname + '_打印单.xlsx'
    wb.save(filename=dest)
    print("保存成功！文件位置:", dest)
    # 执行windows命令，打开excel
    # os.system(r'start' + dest)
    # 执行macOS命令，打开excel
    os.system(r'open ' + dest)

    print('---修改后---')
    st.printWB(ws)


if __name__ == '__main__':
    # print(sys.argv)
    if len(sys.argv) >= 2:
        print(sys.argv)
        sourcefile = sys.argv[1]
    else:
        sourcefile = root_path + '/resources/source.csv'
    process(sourcefile)