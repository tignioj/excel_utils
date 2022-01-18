import os
import sys
from pathlib import Path
from openpyxl.utils import get_column_letter
# libpath = sys.path[0]
from var_val import root_path

sys.path.append(root_path)
from myutil.sheet_utils import sheetutil as st
from myutil.time_utils import getDay
from myutil.file_utils import getParentFolder, getFileName
from openpyxl.styles import Font, Fill, Border, Side


def process(sourcefile):
    # 打开一张表
    wb = st.getWorkBook(sourcefile)
    ws = wb.active
    # st.printWB(ws,2)
    # 删掉订购量为0的行
    st.delete_rows_when_equals(ws, '订购量', ['0'])
    # 删掉不必要的列
    # st.deleteByColumnNames(ws,['商品编码','序号','商品简码','需求量','批零差'])
    st.deleteByColumnNames(ws, ['商品编码', '序号', '商品简码', '需求量'])
    st.printWB(ws)

    # 调整样式
    # 行高
    for i in range(ws.max_row):
        ws.row_dimensions[i + 1].height = 22.2

    # 样式-边
    thin = Side(border_style="thin", color="000000")
    b = Border(top=thin, left=thin, right=thin, bottom=thin)
    f = Font(name='Arial', size=16)
    for row in ws.rows:
        for cell in row:
            cell.font = f
            cell.border = b

    # 给最后一行添加日期
    ws.cell(ws.max_row, 1).value = getDay()

    # 改列名
    st.rename_firstcolumn(ws, '订购量', '订')
    st.rename_firstcolumn(ws, '零售指导价', '指导价')

    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '商品名称')].width = 28
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '批发价')].width = 12
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '指导价')].width = 12
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '订')].width = 4.2
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '金额')].width = 14
    ws.column_dimensions[st.getColumnLetterByColumnName(ws, '批零差')].width = 13

    # 保存
    if len(sys.argv) >= 2:
        save_folder = getParentFolder(sys.argv[1]) + "/printer"
    else:
        save_folder = root_path + '/printer'

    if not os.path.isdir(save_folder):
        os.mkdir(save_folder)
    fname = getFileName(sourcefile)
    wb.save(filename=save_folder + '/' + getDay() + "_" + fname + '_打印单.xlsx')


if __name__ == '__main__':
    # print(sys.argv)
    if len(sys.argv) >= 2:
        print(sys.argv)
        sourcefile = sys.argv[1]
    else:
        sourcefile = root_path + '/resources/source.xls'
    process(sourcefile)
