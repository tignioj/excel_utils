import openpyxl
from xls2xlsx import XLS2XLSX
from myutil.file_utils import getFileExtension
from openpyxl.utils import get_column_letter

class sheetutil:
    @staticmethod
    def xls2xlsx(source):
        # 转换表
        x2x = XLS2XLSX(source)
        # xlsx_fname = getTodayYearMonthDayHourMinSec() + "_converted.xlsx"
        xlsx_fname = source + ".xlsx"
        x2x.to_xlsx(xlsx_fname)
        wb = openpyxl.load_workbook(xlsx_fname)
        return wb
    @staticmethod
    def getColumnLetterByColumnName(ws, colname):
        index = sheetutil.getIndexByColumnName(ws, colname)
        return get_column_letter(index)


    @staticmethod
    def getWorkBook(fname=""):
        if type(fname) is not str:
            raise Exception('file name error!')
        ext = getFileExtension(fname)
        if str(ext).lower() == ".xls":
            return sheetutil.xls2xlsx(fname)
        else:
            return openpyxl.load_workbook(fname)


    @staticmethod
    def printWB(ws=None, maxrow=None):
        print(ws.max_row)
        rows = tuple(ws.iter_rows(max_row=maxrow, values_only=True))
        for row in rows:
            print(row)

    @staticmethod
    def deleteByColumnNames(ws=None, names=None):
        """
        根据首行某列名称删掉该列
        :param names: 要删掉的列表
        :return:
        """
        if type(names) is not list:
            return

        for name in names:
            index = sheetutil.getIndexByColumnName(ws,name)
            ws.delete_cols(index)

    @staticmethod
    def getIndexByColumnName(ws,name):
        """
        根据列名称获取下标，例如('姓名', '年龄', '性别'), 在这里我要获取年龄的下标，就是2
        :param name: 要获取的列名称
        :return: 返回下标，从1开始
        """
        rows = tuple(ws.values)
        if name is None or len(rows) <= 0:
            res = -1
            return res
        first_line = rows[0]
        index = 0
        for c in first_line:
            index += 1
            if c.strip() == name.strip():
                return index
        return -1

    # 根据条件删除Excel
    # https://stackoverflow.com/questions/52821618/openpyxl-how-to-delete-rows-from-an-excel-file-based-on-some-condition/52822133
    @staticmethod
    def delete_rows_when_equals(ws, column, condictions):
        """
        根据条件删除
        :param ws:
        :param columns:
        :param condictions:
        :return:
        """
        index = sheetutil.getIndexByColumnName(ws,column)
        i = 1
        while i <= ws.max_row:
            if ws.cell(row=i, column=index).value in condictions:
                ws.delete_rows(i, 1)
            else: i+=1

    @staticmethod
    def get_data(ws):
        return ws.values

    @staticmethod
    def rename_firstcolumn(ws,before_col, after_col):
        """
        更改第一行的值
        :param before_col:
        :param after_col:
        :return:
        """
        index = sheetutil.getIndexByColumnName(ws,before_col)
        ws.cell(1, index).value = after_col

    @staticmethod
    def append(ws,columns, values):
        ws.append(columns)
        for v in values:
            ws.append(v)
    @staticmethod
    def appendSpec(ws, columns_spec, values):
        """
        指定列添加数据，例如 ['A', 'C', 'D'], [['数据1, '数据2', '数据3'][..]]
        表面数据添加到A、C、D列。
        :param ws:
        :param columns_spec:
        :param values:
        :return:
        """
        mr = ws.max_row
        for v in values:
            index = 0
            for c in columns_spec:
                position = c + str(mr)
                ws[position] = v[index]
                mr+=1
                index +=1
